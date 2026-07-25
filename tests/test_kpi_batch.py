"""
Tests for the local KPI batch harness (src/evaluation/kpi_batch.py).

All tests use mocked pipeline callables — no real indexing or LLM calls.
"""
from __future__ import annotations

import json
from pathlib import Path
from types import SimpleNamespace

import pytest

from openpyxl import load_workbook

from src.evaluation.kpi_batch import (
    EXCEL_HEADERS,
    KpiRow,
    KpiSpec,
    PdfKpiPair,
    ResultRow,
    discover_pairs,
    extract_row,
    iter_active_rows,
    load_kpi_spec,
    run_batch,
    shorten_quote,
    write_workbook,
)


# ---------------------------------------------------------------------------
# Helpers / fixtures
# ---------------------------------------------------------------------------


def _make_pdf(path: Path, content: bytes = b"%PDF-1.4 fake\n") -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(content)
    return path


def _make_json(path: Path, payload: dict) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload), encoding="utf-8")
    return path


def _default_spec_payload() -> dict:
    return {
        "excel_table": "template.xlsx",
        "excel_sheet": "template",
        "start_row": 1,
        "start_col": 1,
        "KPIs": {
            "1": "Label one",
            "2": "",                 # empty -> falls back to query text
            "3": "Label three",
            "4": "Label four",       # inactive
            "5": "Label five",       # empty query -> skipped
        },
        "queries": {
            "1": "Question one?",
            "2": "Question two?",
            "3": "Question three?",
            "4": "Question four?",
            "5": "   ",              # whitespace-only -> skipped
        },
        "templates": {"1": "template body"},
        "templates_mapping": {
            "1": "1", "2": "1", "3": "1", "4": "1", "5": "1",
        },
        "isActive": {
            "1": True,
            "2": True,
            "3": True,
            "4": False,              # inactive -> skipped
            "5": True,
        },
    }


def _make_citation(quote_text: str, page_number: int, status: str = "valid"):
    return SimpleNamespace(
        quote_text=quote_text,
        page_number=page_number,
        validation_status=status,
    )


def _make_response(answer_text: str, sources=None):
    return SimpleNamespace(answer_text=answer_text, sources=sources or [])


# ---------------------------------------------------------------------------
# discover_pairs
# ---------------------------------------------------------------------------


class TestDiscoverPairs:
    def test_pairs_by_exact_stem(self, tmp_path):
        pdf_dir = tmp_path / "pdfs"
        json_dir = tmp_path / "jsons"
        _make_pdf(pdf_dir / "alpha.pdf")
        _make_pdf(pdf_dir / "beta.pdf")
        _make_json(json_dir / "alpha.json", _default_spec_payload())
        _make_json(json_dir / "beta.json", _default_spec_payload())

        pairs = discover_pairs(pdf_dir, json_dir)
        assert len(pairs) == 2
        assert {p.stem for p in pairs} == {"alpha", "beta"}
        assert all(isinstance(p, PdfKpiPair) for p in pairs)

    def test_pdf_without_matching_json_is_skipped(self, tmp_path):
        pdf_dir = tmp_path / "pdfs"
        json_dir = tmp_path / "jsons"
        _make_pdf(pdf_dir / "orphan.pdf")
        _make_pdf(pdf_dir / "kept.pdf")
        _make_json(json_dir / "kept.json", _default_spec_payload())

        pairs = discover_pairs(pdf_dir, json_dir)
        assert [p.stem for p in pairs] == ["kept"]

    def test_returns_empty_when_no_pdfs(self, tmp_path):
        pdf_dir = tmp_path / "pdfs"
        json_dir = tmp_path / "jsons"
        pdf_dir.mkdir()
        json_dir.mkdir()
        assert discover_pairs(pdf_dir, json_dir) == []


# ---------------------------------------------------------------------------
# load_kpi_spec / iter_active_rows
# ---------------------------------------------------------------------------


class TestLoadKpiSpec:
    def test_skips_inactive_rows(self, tmp_path):
        path = _make_json(tmp_path / "spec.json", _default_spec_payload())
        spec = load_kpi_spec(path)
        ids = [r.kpi_id for r in spec.rows]
        assert "4" not in ids  # inactive

    def test_skips_empty_queries(self, tmp_path):
        path = _make_json(tmp_path / "spec.json", _default_spec_payload())
        spec = load_kpi_spec(path)
        ids = [r.kpi_id for r in spec.rows]
        assert "5" not in ids  # whitespace-only query

    def test_empty_label_falls_back_to_query(self, tmp_path):
        path = _make_json(tmp_path / "spec.json", _default_spec_payload())
        spec = load_kpi_spec(path)
        row_two = next(r for r in spec.rows if r.kpi_id == "2")
        assert row_two.kpi_label == row_two.query == "Question two?"

    def test_keeps_template_id(self, tmp_path):
        path = _make_json(tmp_path / "spec.json", _default_spec_payload())
        spec = load_kpi_spec(path)
        assert all(r.template_id == "1" for r in spec.rows)

    def test_iter_active_rows_order(self, tmp_path):
        path = _make_json(tmp_path / "spec.json", _default_spec_payload())
        spec = load_kpi_spec(path)
        ids = [r.kpi_id for r in iter_active_rows(spec)]
        assert ids == ["1", "2", "3"]


# ---------------------------------------------------------------------------
# shorten_quote
# ---------------------------------------------------------------------------


class TestShortenQuote:
    def test_short_text_returned_verbatim(self):
        assert shorten_quote("short") == "short"

    def test_long_text_is_head_substring_plus_ellipsis(self):
        text = "x" * 350
        out = shorten_quote(text, max_len=300)
        assert out.endswith("...")
        assert out[:-3] == text[:300]  # exact substring head
        assert len(out) == 303

    def test_handles_none(self):
        assert shorten_quote(None) == ""  # type: ignore[arg-type]


# ---------------------------------------------------------------------------
# extract_row
# ---------------------------------------------------------------------------


class TestExtractRow:
    def _kpi_row(self) -> KpiRow:
        return KpiRow(kpi_id="1", kpi_label="L1", query="Q1", template_id="1")

    def test_uses_first_citation(self):
        response = _make_response(
            "answer body",
            sources=[
                _make_citation("first quote", 7, "valid"),
                _make_citation("second quote", 12, "unverified"),
            ],
        )
        row = extract_row("doc.pdf", self._kpi_row(), response)
        assert row.summary == "answer body"
        assert row.source_quote == "first quote"
        assert row.page_number == "7"
        assert row.validation_status == "valid"

    def test_no_sources_yields_no_citation_fallback(self):
        response = _make_response("answer body", sources=[])
        row = extract_row("doc.pdf", self._kpi_row(), response)
        assert row.summary == "answer body"
        assert row.source_quote == ""
        assert row.page_number == ""
        assert row.validation_status == "no_citation"

    def test_template_id_carried_through(self):
        response = _make_response("a", sources=[_make_citation("q", 1)])
        kpi = KpiRow(kpi_id="3", kpi_label="L", query="Q", template_id="42")
        row = extract_row("d.pdf", kpi, response)
        assert row.template_id == "42"

    def test_long_quote_is_shortened(self):
        long_quote = "y" * 350
        response = _make_response("a", sources=[_make_citation(long_quote, 1)])
        row = extract_row("d.pdf", self._kpi_row(), response)
        assert row.source_quote.endswith("...")
        assert len(row.source_quote) == 303


# ---------------------------------------------------------------------------
# write_workbook
# ---------------------------------------------------------------------------


class TestWriteWorkbook:
    def _row(self, **overrides) -> ResultRow:
        defaults = dict(
            pdf_name="x.pdf",
            kpi_id="1",
            kpi_label="L",
            query="Q",
            summary="S",
            source_quote="q",
            page_number="3",
            validation_status="valid",
            template_id="1",
            timestamp="2026-01-01T00:00:00Z",
        )
        defaults.update(overrides)
        return ResultRow(**defaults)

    def test_creates_file(self, tmp_path):
        out = write_workbook([self._row()], tmp_path / "out.xlsx")
        assert out.exists()

    def test_header_row(self, tmp_path):
        write_workbook([self._row()], tmp_path / "out.xlsx")
        wb = load_workbook(tmp_path / "out.xlsx")
        ws = wb["results"]
        headers = [c.value for c in ws[1]]
        assert headers == EXCEL_HEADERS

    def test_data_row_round_trip(self, tmp_path):
        write_workbook(
            [self._row(summary="hello", page_number="9")],
            tmp_path / "out.xlsx",
        )
        wb = load_workbook(tmp_path / "out.xlsx")
        ws = wb["results"]
        values = [c.value for c in ws[2]]
        assert values[0] == "x.pdf"
        assert values[4] == "hello"
        assert values[6] == "9"


# ---------------------------------------------------------------------------
# run_batch
# ---------------------------------------------------------------------------


class TestRunBatch:
    def _setup(self, tmp_path):
        pdf_dir = tmp_path / "pdfs"
        json_dir = tmp_path / "jsons"
        out_dir = tmp_path / "out"
        _make_pdf(pdf_dir / "alpha.pdf")
        _make_json(json_dir / "alpha.json", _default_spec_payload())
        return pdf_dir, json_dir, out_dir

    def test_writes_one_workbook_per_pdf(self, tmp_path):
        pdf_dir, json_dir, out_dir = self._setup(tmp_path)

        def fake_index(pdf_path, *, index_dir, **kwargs):
            return SimpleNamespace(ok=True)

        def fake_answer(query, *, index_dir, **kwargs):
            return _make_response(
                f"answer for {query}",
                sources=[_make_citation("q-text", 2)],
            )

        written = run_batch(
            pdf_dir, json_dir, out_dir,
            index_doc=fake_index, answer=fake_answer,
        )
        assert len(written) == 1
        assert written[0].name == "alpha_results.xlsx"
        assert written[0].exists()

    def test_rows_written_match_active_kpis(self, tmp_path):
        pdf_dir, json_dir, out_dir = self._setup(tmp_path)

        def fake_index(pdf_path, *, index_dir, **kwargs):
            return None

        def fake_answer(query, *, index_dir, **kwargs):
            return _make_response("a", sources=[_make_citation("q", 1)])

        run_batch(
            pdf_dir, json_dir, out_dir,
            index_doc=fake_index, answer=fake_answer,
        )
        wb = load_workbook(out_dir / "alpha_results.xlsx")
        ws = wb["results"]
        # header + 3 active rows (ids 1, 2, 3)
        assert ws.max_row == 4

    def test_query_failure_recorded_as_error_row(self, tmp_path):
        pdf_dir, json_dir, out_dir = self._setup(tmp_path)

        def fake_index(pdf_path, *, index_dir, **kwargs):
            return None

        def fake_answer(query, *, index_dir, **kwargs):
            raise RuntimeError("boom")

        run_batch(
            pdf_dir, json_dir, out_dir,
            index_doc=fake_index, answer=fake_answer,
        )
        wb = load_workbook(out_dir / "alpha_results.xlsx")
        ws = wb["results"]
        # All data rows are error rows.
        summary_col = EXCEL_HEADERS.index("summary") + 1
        status_col = EXCEL_HEADERS.index("validation_status") + 1
        for row_idx in range(2, ws.max_row + 1):
            assert ws.cell(row=row_idx, column=summary_col).value.startswith("ERROR: ")
            assert ws.cell(row=row_idx, column=status_col).value == "error"

    def test_index_failure_short_circuits_pdf(self, tmp_path):
        pdf_dir, json_dir, out_dir = self._setup(tmp_path)

        def fake_index(pdf_path, *, index_dir, **kwargs):
            raise RuntimeError("indexing failed")

        called = []

        def fake_answer(query, *, index_dir, **kwargs):
            called.append(query)
            return _make_response("a")

        run_batch(
            pdf_dir, json_dir, out_dir,
            index_doc=fake_index, answer=fake_answer,
        )
        assert called == []  # answer was never invoked
        wb = load_workbook(out_dir / "alpha_results.xlsx")
        ws = wb["results"]
        status_col = EXCEL_HEADERS.index("validation_status") + 1
        for row_idx in range(2, ws.max_row + 1):
            assert ws.cell(row=row_idx, column=status_col).value == "error"

    def test_unpaired_pdf_is_skipped(self, tmp_path):
        pdf_dir = tmp_path / "pdfs"
        json_dir = tmp_path / "jsons"
        out_dir = tmp_path / "out"
        _make_pdf(pdf_dir / "orphan.pdf")
        json_dir.mkdir()

        written = run_batch(
            pdf_dir, json_dir, out_dir,
            index_doc=lambda *a, **k: None,
            answer=lambda *a, **k: _make_response("a"),
        )
        assert written == []

    def test_no_citation_row_uses_fallback(self, tmp_path):
        pdf_dir, json_dir, out_dir = self._setup(tmp_path)

        def fake_index(*a, **k):
            return None

        def fake_answer(query, *, index_dir, **kwargs):
            return _make_response("a", sources=[])

        run_batch(
            pdf_dir, json_dir, out_dir,
            index_doc=fake_index, answer=fake_answer,
        )
        wb = load_workbook(out_dir / "alpha_results.xlsx")
        ws = wb["results"]
        status_col = EXCEL_HEADERS.index("validation_status") + 1
        for row_idx in range(2, ws.max_row + 1):
            assert ws.cell(row=row_idx, column=status_col).value == "no_citation"
